import io
import os
import re
import time
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path

import pandas as pd
import streamlit as st
from openpyxl import load_workbook


PRODUCT_TERMS = ["Suncron", "Sunfix", "Sunzol", "Suncion", "APEX", "SP-SE", "CP-R", "CP-D"]
MAX_RESIDUE_TRANSLATION_PASSES = 3
APP_PASSWORD = "5050"


MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
CT_NS = "http://schemas.openxmlformats.org/package/2006/content-types"


def _merge_sheet_drawing_nodes(original_sheet_xml: bytes, translated_sheet_xml: bytes) -> bytes:
	"""Keep drawing anchors (image references) from the original worksheet XML."""
	try:
		original_root = ET.fromstring(original_sheet_xml)
		translated_root = ET.fromstring(translated_sheet_xml)
	except ET.ParseError:
		return translated_sheet_xml

	drawing_tags = ["drawing", "legacyDrawing", "legacyDrawingHF", "picture"]

	for tag in drawing_tags:
		qtag = f"{{{MAIN_NS}}}{tag}"
		for existing in list(translated_root.findall(qtag)):
			translated_root.remove(existing)
		for source_node in original_root.findall(qtag):
			translated_root.append(ET.fromstring(ET.tostring(source_node)))

	ET.register_namespace("", MAIN_NS)
	return ET.tostring(translated_root, encoding="utf-8", xml_declaration=True)


def _merge_content_types(original_content_types_xml: bytes, translated_content_types_xml: bytes) -> bytes:
	"""Ensure drawing/image content-types survive after openpyxl save."""
	try:
		original_root = ET.fromstring(original_content_types_xml)
		translated_root = ET.fromstring(translated_content_types_xml)
	except ET.ParseError:
		return translated_content_types_xml

	existing_keys = set()
	for child in translated_root:
		if child.tag.endswith("Default"):
			existing_keys.add(("Default", child.attrib.get("Extension", "").lower()))
		elif child.tag.endswith("Override"):
			existing_keys.add(("Override", child.attrib.get("PartName", "")))

	for child in original_root:
		if child.tag.endswith("Default"):
			key = ("Default", child.attrib.get("Extension", "").lower())
		elif child.tag.endswith("Override"):
			key = ("Override", child.attrib.get("PartName", ""))
		else:
			continue

		if key not in existing_keys:
			translated_root.append(ET.fromstring(ET.tostring(child)))
			existing_keys.add(key)

	ET.register_namespace("", CT_NS)
	return ET.tostring(translated_root, encoding="utf-8", xml_declaration=True)


def _restore_embedded_graphics(original_xlsx: bytes, translated_xlsx: bytes) -> bytes:
	"""Restore embedded drawings/images dropped by openpyxl round-trip."""
	if not original_xlsx or not translated_xlsx:
		return translated_xlsx

	try:
		with zipfile.ZipFile(io.BytesIO(original_xlsx), "r") as z_orig, zipfile.ZipFile(
			io.BytesIO(translated_xlsx), "r"
		) as z_new:
			original_names = set(z_orig.namelist())
			merged_files = {name: z_new.read(name) for name in z_new.namelist()}

			sheet_rels_with_drawing = []
			for name in original_names:
				if not (name.startswith("xl/worksheets/_rels/") and name.endswith(".rels")):
					continue
				rels_xml = z_orig.read(name)
				if b"/drawing" in rels_xml or b"legacyDrawing" in rels_xml:
					sheet_rels_with_drawing.append(name)
					merged_files[name] = rels_xml

			for name in original_names:
				if name.startswith("xl/drawings/") or name.startswith("xl/media/"):
					merged_files[name] = z_orig.read(name)

			for rel_name in sheet_rels_with_drawing:
				sheet_file = rel_name.replace("xl/worksheets/_rels/", "xl/worksheets/").replace(".rels", "")
				if sheet_file in merged_files and sheet_file in original_names:
					merged_files[sheet_file] = _merge_sheet_drawing_nodes(
						z_orig.read(sheet_file),
						merged_files[sheet_file],
					)

			if "[Content_Types].xml" in original_names and "[Content_Types].xml" in merged_files:
				merged_files["[Content_Types].xml"] = _merge_content_types(
					z_orig.read("[Content_Types].xml"),
					merged_files["[Content_Types].xml"],
				)

			result_buffer = io.BytesIO()
			with zipfile.ZipFile(result_buffer, "w", compression=zipfile.ZIP_DEFLATED) as z_out:
				for name, payload in merged_files.items():
					z_out.writestr(name, payload)

			return result_buffer.getvalue()
	except Exception:
		# If anything goes wrong, return the translated workbook without blocking download.
		return translated_xlsx


# 1. 페이지 설정
st.set_page_config(page_title="MSDS 자동 번역 & 엑셀 생성 앱", layout="wide")
st.title("🌐 MSDS 다국어 자동 번역기 (서식 보존형)")
st.caption("영문 MSDS(_ENG) 엑셀을 업로드하면, 서식을 유지한 채 지정한 언어로 번역된 엑셀 파일을 생성합니다.")


def _normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
	df = df.copy()
	df.columns = [str(col).strip() for col in df.columns]
	return df


def _normalize_text_for_lookup(text: str) -> str:
	# Unify whitespace and temperature symbols so dictionary lookup is less fragile.
	text = format_temperature(str(text))
	text = re.sub(r"\s+", " ", text).strip()
	return text


def _build_translation_dict(df: pd.DataFrame, target_lang: str) -> dict:
	if "English" not in df.columns or target_lang not in df.columns:
		return {}

	valid_df = df.dropna(subset=["English", target_lang]).copy()
	valid_df["English"] = valid_df["English"].astype(str).map(_normalize_text_for_lookup)
	valid_df[target_lang] = valid_df[target_lang].astype(str).str.strip()
	valid_df = valid_df[(valid_df["English"] != "") & (valid_df[target_lang] != "")]

	return dict(zip(valid_df["English"], valid_df[target_lang]))


# 2. 마스터 DB 불러오기 및 사전(Dictionary)화
@st.cache_data
def load_translation_dict(target_lang: str):
	"""
	우선순위:
	1) 로컬 Master_DB.xlsx
	2) 로컬 master_db.csv

	필요 컬럼: [ English | Vietnamese | Spanish | Chinese ]
	"""
	trans_dict: dict[str, str] = {}
	source_logs: list[str] = []

	# 1순위: 로컬 Master_DB.xlsx (고정)
	local_master_xlsx = Path("Master_DB.xlsx")
	if local_master_xlsx.exists():
		try:
			local_xlsx_df = pd.read_excel(local_master_xlsx, dtype=str)
			local_xlsx_df = _normalize_columns(local_xlsx_df)
			local_dict = _build_translation_dict(local_xlsx_df, target_lang)
			added_count = 0
			for eng, tr in local_dict.items():
				if eng not in trans_dict:
					trans_dict[eng] = tr
					added_count += 1
			source_logs.append(
				f"로컬 Master_DB.xlsx에서 {added_count:,}개 번역 매핑 로드 (1순위)"
			)
		except Exception as exc:  # noqa: BLE001
			source_logs.append(f"로컬 Master_DB.xlsx 로드 실패: {exc}")
	else:
		source_logs.append("로컬 Master_DB.xlsx 파일 없음")

	# 2순위: 로컬 master_db.csv (백업)
	local_master_csv = Path("master_db.csv")
	if local_master_csv.exists():
		try:
			csv_df = pd.read_csv(local_master_csv, dtype=str)
			csv_df = _normalize_columns(csv_df)
			csv_dict = _build_translation_dict(csv_df, target_lang)
			added_count = 0
			for eng, tr in csv_dict.items():
				if eng not in trans_dict:
					trans_dict[eng] = tr
					added_count += 1
			source_logs.append(
				f"로컬 master_db.csv에서 {added_count:,}개 번역 매핑 추가 (2순위, 중복 키 제외)"
			)
		except Exception as exc:  # noqa: BLE001
			source_logs.append(f"로컬 master_db.csv 로드 실패: {exc}")
	else:
		source_logs.append("로컬 master_db.csv 파일 없음")

	if not trans_dict:
		st.error("번역 사전을 생성하지 못했습니다. Master_DB.xlsx 또는 master_db.csv 형식을 확인해주세요.")

	return trans_dict, source_logs


# 3. 핵심 번역 예외 처리 규칙
def format_temperature(text):
	if not isinstance(text, str):
		return text
	# Normalize numeric temperature notations such as 25oC, 270.C, 80 ° C.
	text = re.sub(r"(?i)(\d+(?:\.\d+)?)\s*[°ºo]?\s*\.?\s*c\b", r"\1℃", text)
	text = re.sub(r"°\s*C|º\s*C|℃", "℃", text, flags=re.IGNORECASE)
	text = re.sub(r"\bdeg(?:ree)?s?\s*C\b", "℃", text, flags=re.IGNORECASE)
	text = re.sub(r"\bdegrees?\s+c\b", "℃", text, flags=re.IGNORECASE)
	text = re.sub(r"\bgrados?\s+c(?:elsius)?\b", "℃", text, flags=re.IGNORECASE)
	text = re.sub(r"\b摄氏度\b|\b攝氏度\b", "℃", text)
	text = re.sub(r"\bcelsius\b|\bcentigrade\b", "℃", text, flags=re.IGNORECASE)
	# Force requested display format: 270 ℃, 25 ℃.
	text = re.sub(r"(\d+(?:\.\d+)?)\s*℃", r"\1 ℃", text)
	return text


def should_force_keep_product_phrase(text: str) -> bool:
	if not isinstance(text, str):
		return False

	brand_pat = r"\b(" + "|".join(re.escape(term) for term in PRODUCT_TERMS) + r")\b"
	code_pat = r"\b(CP-[A-Z]|SP-SE)\b"

	# If a line looks like a product-name value (brand + product code), keep it unchanged.
	if re.search(brand_pat, text, re.IGNORECASE) and re.search(code_pat, text, re.IGNORECASE):
		return True

	return False


def should_keep_original(text):
	if not isinstance(text, str):
		return True

	# 규칙 1: 제품명 보호 (문장 전체가 제품명/코드일 때만 보호)
	product_pat = r"\b(" + "|".join(re.escape(term) for term in PRODUCT_TERMS) + r")\b"
	if re.search(product_pat, text, re.IGNORECASE):
		if len(text.strip()) <= 60 and re.fullmatch(r"[A-Za-z0-9\-\s_/]+", text.strip()):
			return True

	# 규칙 2: 복잡한 화학물질명 보호 (과도한 보호 방지를 위해 CAS 형식 중심으로 제한)
	if re.search(r"\b\d{2,7}-\d{2}-\d\b", text):
		return True
	if re.search(r"C\.I\.\s*(disperse|reactive|acid|direct|basic|vat)", text, re.IGNORECASE):
		return True

	chem_keywords = [
		"lignosulfonate",
		"Naphthalenesulfonic acid",
		"polymer with formaldehyde",
		"methyl ester",
		"methoxyphenyl",
		"Glycine",
	]
	if any(keyword.lower() in text.lower() for keyword in chem_keywords):
		return True

	return False


def _build_phrase_rules(trans_dict: dict[str, str]):
	# Keep partial replacement conservative to avoid mistranslating tiny tokens like "Red".
	rules = []
	for source, target in trans_dict.items():
		src = str(source).strip()
		if not src or not target:
			continue

		norm = _normalize_text_for_lookup(src)
		word_count = len(norm.split())
		if len(norm) < 12:
			continue
		if word_count == 1 and len(norm) < 20:
			continue

		pattern = re.escape(src)
		if re.match(r"^[A-Za-z0-9]", src):
			pattern = r"(?<![A-Za-z0-9])" + pattern
		if re.search(r"[A-Za-z0-9]$", src):
			pattern = pattern + r"(?![A-Za-z0-9])"

		try:
			rules.append((len(norm), re.compile(pattern, re.IGNORECASE), str(target)))
		except re.error:
			continue

	# Longer phrases first for stable replacement.
	rules.sort(key=lambda x: x[0], reverse=True)
	return rules


def _get_msds_glossary(target_lang: str):
	# Fallback terminology for common SDS headings and regulatory phrases.
	common_en = {
		"safety data sheet": {
			"Vietnamese": "phiếu an toàn hóa chất",
			"Spanish": "ficha de datos de seguridad",
			"Chinese": "化学品安全技术说明书",
		},
		"according to": {"Vietnamese": "theo", "Spanish": "según", "Chinese": "根据"},
		"revision date": {
			"Vietnamese": "ngày sửa đổi",
			"Spanish": "fecha de revisión",
			"Chinese": "修订日期",
		},
		"version": {"Vietnamese": "phiên bản", "Spanish": "versión", "Chinese": "版本"},
		"product name": {"Vietnamese": "tên sản phẩm", "Spanish": "nombre del producto", "Chinese": "产品名称"},
		"product": {"Vietnamese": "sản phẩm", "Spanish": "producto", "Chinese": "产品"},
		"recommended use": {
			"Vietnamese": "mục đích sử dụng khuyến nghị",
			"Spanish": "uso recomendado",
			"Chinese": "推荐用途",
		},
		"supplier": {"Vietnamese": "nhà cung cấp", "Spanish": "proveedor", "Chinese": "供应商"},
		"manufacturer": {"Vietnamese": "nhà sản xuất", "Spanish": "fabricante", "Chinese": "制造商"},
		"emergency telephone number": {
			"Vietnamese": "số điện thoại khẩn cấp",
			"Spanish": "número de teléfono de emergencia",
			"Chinese": "应急电话",
		},
		"hazard identification": {
			"Vietnamese": "nhận dạng mối nguy",
			"Spanish": "identificación de peligros",
			"Chinese": "危险性概述",
		},
		"composition/information on ingredients": {
			"Vietnamese": "thành phần/thông tin về các thành phần",
			"Spanish": "composición/información sobre los componentes",
			"Chinese": "成分/组成信息",
		},
		"first aid measures": {
			"Vietnamese": "biện pháp sơ cứu",
			"Spanish": "medidas de primeros auxilios",
			"Chinese": "急救措施",
		},
		"general information": {
			"Vietnamese": "thông tin chung",
			"Spanish": "información general",
			"Chinese": "一般信息",
		},
		"firefighting measures": {
			"Vietnamese": "biện pháp chữa cháy",
			"Spanish": "medidas de lucha contra incendios",
			"Chinese": "消防措施",
		},
		"accidental release measures": {
			"Vietnamese": "biện pháp xử lý sự cố tràn đổ",
			"Spanish": "medidas en caso de vertido accidental",
			"Chinese": "泄漏应急处理",
		},
		"handling and storage": {
			"Vietnamese": "xử lý và bảo quản",
			"Spanish": "manipulación y almacenamiento",
			"Chinese": "操作处置与储存",
		},
		"exposure controls/personal protection": {
			"Vietnamese": "kiểm soát phơi nhiễm/bảo hộ cá nhân",
			"Spanish": "controles de exposición/protección personal",
			"Chinese": "接触控制和个体防护",
		},
		"physical and chemical properties": {
			"Vietnamese": "tính chất vật lý và hóa học",
			"Spanish": "propiedades físicas y químicas",
			"Chinese": "理化特性",
		},
		"stability and reactivity": {
			"Vietnamese": "độ ổn định và khả năng phản ứng",
			"Spanish": "estabilidad y reactividad",
			"Chinese": "稳定性和反应性",
		},
		"toxicological information": {
			"Vietnamese": "thông tin độc tính",
			"Spanish": "información toxicológica",
			"Chinese": "毒理学信息",
		},
		"ecological information": {
			"Vietnamese": "thông tin sinh thái",
			"Spanish": "información ecológica",
			"Chinese": "生态学信息",
		},
		"disposal considerations": {
			"Vietnamese": "cân nhắc khi thải bỏ",
			"Spanish": "consideraciones sobre la eliminación",
			"Chinese": "废弃处置",
		},
		"transport information": {
			"Vietnamese": "thông tin vận chuyển",
			"Spanish": "información sobre el transporte",
			"Chinese": "运输信息",
		},
		"regulatory information": {
			"Vietnamese": "thông tin quy định",
			"Spanish": "información reglamentaria",
			"Chinese": "法规信息",
		},
		"other information": {
			"Vietnamese": "thông tin khác",
			"Spanish": "otra información",
			"Chinese": "其他信息",
		},
		"results of pbt and vpvb assessment": {
			"Vietnamese": "Kết quả đánh giá PBT và vPvB",
			"Spanish": "Resultados de la evaluación PBT y mPmB",
			"Chinese": "PBT和vPvB评估结果",
		},
		"occupational exposure limit values": {
			"Vietnamese": "giá trị giới hạn phơi nhiễm nghề nghiệp",
			"Spanish": "valores límite de exposición ocupacional",
			"Chinese": "职业接触限值",
		},
		"chemical substances in the work environment": {
			"Vietnamese": "các chất hóa học trong môi trường làm việc",
			"Spanish": "sustancias químicas en el ambiente de trabajo",
			"Chinese": "工作环境中的化学物质",
		},
		"limit values for the chemical agents in the air at the working environment": {
			"Vietnamese": "giá trị giới hạn đối với tác nhân hóa học trong không khí tại môi trường làm việc",
			"Spanish": "valores límite para agentes químicos en el aire del ambiente laboral",
			"Chinese": "工作环境空气中化学因子的限值",
		},
		"european union (eu) commission directive": {
			"Vietnamese": "chỉ thị của Ủy ban Liên minh Châu Âu (EU)",
			"Spanish": "directiva de la Comisión de la Unión Europea (UE)",
			"Chinese": "欧盟委员会指令",
		},
		"no general information": {
			"Vietnamese": "không có thông tin chung",
			"Spanish": "sin información general",
			"Chinese": "无一般信息",
		},
		"appearance": {"Vietnamese": "ngoại quan", "Spanish": "aspecto", "Chinese": "外观"},
		"odor": {"Vietnamese": "mùi", "Spanish": "olor", "Chinese": "气味"},
		"skin": {"Vietnamese": "da", "Spanish": "piel", "Chinese": "皮肤"},
		"oral": {"Vietnamese": "đường miệng", "Spanish": "oral", "Chinese": "经口"},
		"dermal": {"Vietnamese": "qua da", "Spanish": "dérmico", "Chinese": "经皮"},
		"inhalation": {"Vietnamese": "hít phải", "Spanish": "inhalación", "Chinese": "吸入"},
		"ph value": {"Vietnamese": "giá trị pH", "Spanish": "valor de pH", "Chinese": "pH值"},
		"boiling point": {"Vietnamese": "điểm sôi", "Spanish": "punto de ebullición", "Chinese": "沸点"},
		"flash point": {"Vietnamese": "điểm chớp cháy", "Spanish": "punto de inflamación", "Chinese": "闪点"},
		"auto-ignition temperature": {
			"Vietnamese": "nhiệt độ tự bốc cháy",
			"Spanish": "temperatura de autoignición",
			"Chinese": "自燃温度",
		},
		"relative density": {"Vietnamese": "khối lượng riêng tương đối", "Spanish": "densidad relativa", "Chinese": "相对密度"},
		"solubility": {"Vietnamese": "độ tan", "Spanish": "solubilidad", "Chinese": "溶解性"},
		"vapor pressure": {"Vietnamese": "áp suất hơi", "Spanish": "presión de vapor", "Chinese": "蒸气压"},
		"vapor density": {"Vietnamese": "khối lượng riêng hơi", "Spanish": "densidad de vapor", "Chinese": "蒸气密度"},
		"not classified": {"Vietnamese": "không được phân loại", "Spanish": "no clasificado", "Chinese": "未分类"},
		"not available": {"Vietnamese": "không có dữ liệu", "Spanish": "no disponible", "Chinese": "无可用数据"},
		"not applicable": {"Vietnamese": "không áp dụng", "Spanish": "no aplicable", "Chinese": "不适用"},
		"wear protective gloves": {
			"Vietnamese": "đeo găng tay bảo hộ",
			"Spanish": "usar guantes de protección",
			"Chinese": "佩戴防护手套",
		},
		"wear protective clothing": {
			"Vietnamese": "mặc quần áo bảo hộ",
			"Spanish": "usar ropa de protección",
			"Chinese": "穿戴防护服",
		},
		"wear eye protection": {
			"Vietnamese": "đeo bảo vệ mắt",
			"Spanish": "usar protección ocular",
			"Chinese": "佩戴护目镜",
		},
		"if inhaled": {"Vietnamese": "nếu hít phải", "Spanish": "si se inhala", "Chinese": "如吸入"},
		"if in eyes": {"Vietnamese": "nếu dính vào mắt", "Spanish": "si entra en los ojos", "Chinese": "如进入眼睛"},
		"if on skin": {"Vietnamese": "nếu dính trên da", "Spanish": "si entra en contacto con la piel", "Chinese": "如沾染皮肤"},
		"if swallowed": {"Vietnamese": "nếu nuốt phải", "Spanish": "si se ingiere", "Chinese": "如误食"},
		"rinse cautiously with water": {
			"Vietnamese": "rửa cẩn thận bằng nước",
			"Spanish": "enjuagar cuidadosamente con agua",
			"Chinese": "用水小心冲洗",
		},
		"seek medical advice": {
			"Vietnamese": "tham khảo ý kiến bác sĩ",
			"Spanish": "consultar a un médico",
			"Chinese": "就医咨询",
		},
		"keep out of reach of children": {
			"Vietnamese": "để xa tầm tay trẻ em",
			"Spanish": "mantener fuera del alcance de los niños",
			"Chinese": "置于儿童接触不到处",
		},
		"store in a well-ventilated place": {
			"Vietnamese": "bảo quản ở nơi thông thoáng",
			"Spanish": "almacenar en un lugar bien ventilado",
			"Chinese": "存放于通风良好处",
		},
		"avoid breathing dust/fume/gas/mist/vapours/spray": {
			"Vietnamese": "tránh hít bụi/khói/khí/sương/hơi/sol khí",
			"Spanish": "evitar respirar polvo/humo/gas/niebla/vapores/aerosoles",
			"Chinese": "避免吸入粉尘/烟雾/气体/薄雾/蒸气/喷雾",
		},
		"causes skin irritation": {
			"Vietnamese": "gây kích ứng da",
			"Spanish": "provoca irritación cutánea",
			"Chinese": "造成皮肤刺激",
		},
		"causes serious eye irritation": {
			"Vietnamese": "gây kích ứng mắt nghiêm trọng",
			"Spanish": "provoca irritación ocular grave",
			"Chinese": "造成严重眼刺激",
		},
		"may cause an allergic skin reaction": {
			"Vietnamese": "có thể gây phản ứng dị ứng da",
			"Spanish": "puede provocar una reacción alérgica en la piel",
			"Chinese": "可能导致皮肤过敏反应",
		},
		"harmful if swallowed": {
			"Vietnamese": "có hại nếu nuốt phải",
			"Spanish": "nocivo en caso de ingestión",
			"Chinese": "吞咽有害",
		},
		"harmful if inhaled": {
			"Vietnamese": "có hại nếu hít phải",
			"Spanish": "nocivo en caso de inhalación",
			"Chinese": "吸入有害",
		},
		"very toxic to aquatic life": {
			"Vietnamese": "rất độc đối với sinh vật thủy sinh",
			"Spanish": "muy tóxico para los organismos acuáticos",
			"Chinese": "对水生生物有剧毒",
		},
		"toxic to aquatic life with long lasting effects": {
			"Vietnamese": "độc đối với sinh vật thủy sinh với tác động kéo dài",
			"Spanish": "tóxico para los organismos acuáticos, con efectos duraderos",
			"Chinese": "对水生生物有毒并具有长期持续影响",
		},
	}

	if target_lang not in {"Vietnamese", "Spanish", "Chinese"}:
		return {}

	return {k: v[target_lang] for k, v in common_en.items() if target_lang in v}


def _get_hardcoded_overrides(target_lang: str):
	# High-impact SDS lines that should be translated consistently when DB matching fails.
	if target_lang == "Vietnamese":
		return {
			"transport in bulk according to annex ii of marpol and the ibc code": "Vận chuyển hàng rời theo Phụ lục II của MARPOL và Bộ luật IBC",
			"european union (eu) transport of dangerous goods by road - dangerous goods list": "Liên minh Châu Âu (EU) Vận chuyển Hàng hóa Nguy hiểm bằng Đường bộ - Danh mục Hàng hóa Nguy hiểm",
			"many factors determine whether the reported hazards are risks in the workplace or other settings. risks may be determined by reference to exposures scenarios. scale of use, frequency of use and current or available engineering controls must be considered.": "Nhiều yếu tố quyết định liệu các mối nguy được nêu có trở thành rủi ro tại nơi làm việc hoặc trong các bối cảnh khác hay không. Rủi ro có thể được xác định dựa trên các kịch bản phơi nhiễm. Cần xem xét quy mô sử dụng, tần suất sử dụng và các biện pháp kiểm soát kỹ thuật hiện có hoặc sẵn có.",
		}
	if target_lang == "Spanish":
		return {
			"transport in bulk according to annex ii of marpol and the ibc code": "Transporte a granel de acuerdo con el Anexo II de MARPOL y el Código IBC",
			"european union (eu) transport of dangerous goods by road - dangerous goods list": "Unión Europea (UE) Transporte de Mercancías Peligrosas por Carretera - Lista de Mercancías Peligrosas",
		}
	if target_lang == "Chinese":
		return {
			"transport in bulk according to annex ii of marpol and the ibc code": "根据MARPOL附则II和IBC规则进行散装运输",
			"european union (eu) transport of dangerous goods by road - dangerous goods list": "欧盟（EU）道路危险货物运输 - 危险货物清单",
		}
	return {}


def _contains_english_text(text: str) -> bool:
	# Treat 3+ consecutive ASCII letters as untranslated English residue.
	return bool(re.search(r"[A-Za-z]{3,}", text or ""))


@st.cache_resource
def _get_online_translator(target_lang: str):
	target_map = {
		"Vietnamese": "vi",
		"Spanish": "es",
		"Chinese": "zh-CN",
	}

	if target_lang not in target_map:
		return None
	try:
		from deep_translator import GoogleTranslator

		return GoogleTranslator(source="en", target=target_map[target_lang])
	except Exception:
		return None


def _mask_product_terms(text: str):
	product_pat = re.compile(r"\b(" + "|".join(re.escape(term) for term in PRODUCT_TERMS) + r")\b", re.IGNORECASE)
	replacements: dict[str, str] = {}
	idx = 0

	def _repl(match):
		nonlocal idx
		key = f"__KEEP_TERM_{idx}__"
		replacements[key] = match.group(0)
		idx += 1
		return key

	return product_pat.sub(_repl, text), replacements


def _unmask_product_terms(text: str, replacements: dict[str, str]):
	result = text
	for key, value in replacements.items():
		result = result.replace(key, value)
	return result


def _preserve_product_terms_from_source(source_text: str, translated_text: str) -> str:
	if not source_text or not translated_text:
		return translated_text

	result = translated_text
	for term in PRODUCT_TERMS:
		pat = re.compile(r"\b" + re.escape(term) + r"\b", re.IGNORECASE)
		source_match = pat.search(source_text)
		if source_match:
			canonical = source_match.group(0)
			result = pat.sub(canonical, result)

	return result


def _finalize_translation(source_text: str, translated_text: str) -> str:
	# Hard guarantees requested by user: keep product names and force temperature symbol.
	result = _preserve_product_terms_from_source(source_text, translated_text)
	result = format_temperature(result)
	# Mandatory replacements must run for all paths, including protected original lines.
	lang = st.session_state.get("target_lang_runtime")
	if isinstance(lang, str) and lang:
		result = _apply_mandatory_term_replacements(result, lang)
	return result


def _apply_mandatory_term_replacements(text: str, target_lang: str) -> str:
	if not isinstance(text, str):
		return text

	mandatory_terms = {
		"Vietnamese": {"product": "Sản phẩm"},
		"Spanish": {"product": "Producto"},
		"Chinese": {"product": "产品"},
	}

	cleanup_patterns = {
		"Vietnamese": [
			(r"(?i)\bnot\s+applicable\b", "Không áp dụng"),
			(r"(?i)\bnot\s*[-_/]?\s*applicable\b", "Không áp dụng"),
			(r"(?i)\bn\s*/\s*a\b", "Không áp dụng"),
		],
		"Spanish": [
			(r"(?i)\bnot\s+applicable\b", "No aplicable"),
			(r"(?i)\bnot\s*[-_/]?\s*applicable\b", "No aplicable"),
			(r"(?i)\bnot\s+aplicable\b", "No aplicable"),
			(r"(?i)\bresults of evaluación\b", "Resultados de la evaluación"),
			(r"(?i)\bn\s*/\s*a\b", "No aplicable"),
		],
		"Chinese": [
			(r"(?i)\bnot\s+applicable\b", "不适用"),
			(r"(?i)\bnot\s*[-_/]?\s*applicable\b", "不适用"),
			(r"(?i)\bn\s*/\s*a\b", "不适用"),
		],
	}

	replacements = mandatory_terms.get(target_lang, {})
	result = text
	for src, dst in replacements.items():
		result = re.sub(r"(?i)\b" + re.escape(src) + r"\b", dst, result)

	for pat, repl in cleanup_patterns.get(target_lang, []):
		result = re.sub(pat, repl, result)
	return result


def _online_translate_precise(text: str, target_lang: str, translator, delay_sec: float = 1.4):
	if target_lang not in {"Vietnamese", "Spanish", "Chinese"} or translator is None:
		return text, False

	if not _contains_english_text(text):
		return text, False

	masked_text, replacements = _mask_product_terms(text)

	def _translate_with_retry(segment: str):
		for _ in range(4):
			try:
				translated_segment = translator.translate(segment)
				if translated_segment and translated_segment.strip():
					return translated_segment.strip()
			except Exception:
				pass
			time.sleep(delay_sec)
		return None

	# Multiline cells are translated line-by-line for better SDS table consistency.
	if "\n" in masked_text:
		lines = masked_text.splitlines()
		out_lines = []
		changed = False
		for line in lines:
			if not line.strip() or not _contains_english_text(line):
				out_lines.append(line)
				continue
			translated_line = _translate_with_retry(line)
			if translated_line:
				out_lines.append(translated_line)
				changed = True
			else:
				out_lines.append(line)
			time.sleep(delay_sec)

		result = "\n".join(out_lines)
		result = _unmask_product_terms(result, replacements)
		result = _finalize_translation(text, result)
		return result, changed

	if len(masked_text) > 240 and re.search(r"[.!?]", masked_text):
		parts = re.split(r"(?<=[.!?])\s+", masked_text)
		out_parts = []
		changed = False
		for part in parts:
			if not _contains_english_text(part):
				out_parts.append(part)
				continue
			part_translated = _translate_with_retry(part)
			if part_translated:
				out_parts.append(part_translated)
				changed = True
			else:
				out_parts.append(part)
			time.sleep(delay_sec)
		if changed:
			result = " ".join(out_parts)
			result = _unmask_product_terms(result, replacements)
			result = _finalize_translation(text, result)
			return result, True

	translated = _translate_with_retry(masked_text)
	time.sleep(delay_sec)
	if translated:
		result = _unmask_product_terms(translated, replacements)
		result = _finalize_translation(text, result)
		return result, True

	return text, False


def _translate_english_residue(text: str, target_lang: str, translator, glossary_rules, delay_sec: float = 1.6):
	"""Last-chance pass: translate remaining English fragments when any residue is detected."""
	if target_lang not in {"Vietnamese", "Spanish", "Chinese"} or translator is None:
		return text, False
	if not _contains_english_text(text):
		return text, False

	# Apply glossary once more before expensive translation calls.
	pre_glossary_text, _ = _apply_glossary_rules(text, glossary_rules)
	if not _contains_english_text(pre_glossary_text):
		return pre_glossary_text, True

	masked_text, replacements = _mask_product_terms(pre_glossary_text)

	def _try_translate(fragment: str):
		for _ in range(4):
			try:
				result = translator.translate(fragment)
				if result and result.strip():
					return result.strip()
			except Exception:
				pass
			time.sleep(delay_sec)
		return None

	# Split by line first for SDS tables/paragraphs.
	lines = masked_text.splitlines() if "\n" in masked_text else [masked_text]
	out_lines = []
	changed = False

	for line in lines:
		if not line.strip() or not _contains_english_text(line):
			out_lines.append(line)
			continue

		translated_line = _try_translate(line)
		if translated_line:
			out_lines.append(translated_line)
			changed = True
		else:
			# Fragment fallback for stubborn lines: sentence-level attempts.
			parts = re.split(r"(?<=[.!?;:])\s+", line)
			part_out = []
			part_changed = False
			for part in parts:
				if not _contains_english_text(part):
					part_out.append(part)
					continue
				tr = _try_translate(part)
				if tr:
					part_out.append(tr)
					part_changed = True
				else:
					part_out.append(part)
				time.sleep(delay_sec)
			merged = " ".join(part_out)
			out_lines.append(merged)
			if part_changed:
				changed = True

		time.sleep(delay_sec)

	result = "\n".join(out_lines) if "\n" in masked_text else out_lines[0]
	result = _unmask_product_terms(result, replacements)
	result = _finalize_translation(text, result)
	return result, changed


def _build_glossary_rules(target_lang: str):
	glossary = _get_msds_glossary(target_lang)
	rules = []
	for src, dst in glossary.items():
		if not src.strip():
			continue
		if len(src) <= 4 and " " not in src:
			pattern = r"\b" + re.escape(src) + r"\b"
		else:
			pattern = re.escape(src)
			if re.search(r"[A-Za-z]", src):
				pattern = r"(?<![A-Za-z0-9])" + pattern + r"(?![A-Za-z0-9])"
		rules.append((len(src), re.compile(pattern, re.IGNORECASE), dst))
	rules.sort(key=lambda x: x[0], reverse=True)
	return rules


def _translate_msds_templates(text: str, target_lang: str):
	if target_lang not in {"Vietnamese", "Spanish", "Chinese"}:
		return text, False

	section_prefix = {
		"Vietnamese": "PHẦN",
		"Spanish": "SECCIÓN",
		"Chinese": "第",
	}

	signal_words = {
		"Vietnamese": {"danger": "Nguy hiem", "warning": "Canh bao"},
		"Spanish": {"danger": "Peligro", "warning": "Atencion"},
		"Chinese": {"danger": "危险", "warning": "警告"},
	}

	result = text
	changed = False
	glossary = _get_msds_glossary(target_lang)

	def _is_all_caps_ascii(value: str) -> bool:
		letters = re.findall(r"[A-Za-z]", value)
		return bool(letters) and all(ch.isupper() for ch in letters)

	# Standard SDS section heading fallback.
	def section_repl(match):
		nonlocal changed
		changed = True
		num = match.group(1)
		title = match.group(2).strip()
		title_key = _normalize_text_for_lookup(title).lower()
		translated_title = glossary.get(title_key, title)
		if target_lang != "Chinese" and _is_all_caps_ascii(title):
			translated_title = translated_title.upper()
		if target_lang == "Chinese":
			return f"{section_prefix[target_lang]}{num}节: {translated_title}"
		return f"{section_prefix[target_lang]} {num}: {translated_title}"

	result = re.sub(r"(?i)\bsection\s*(\d{1,2})\s*[:\.-]\s*([^\n]+)", section_repl, result)

	# Standard signal words (GHS aligned per language).
	for pat, repl in [
		(r"(?i)\bdanger\b", signal_words[target_lang]["danger"]),
		(r"(?i)\bwarning\b", signal_words[target_lang]["warning"]),
	]:
		new_result, n = re.subn(pat, repl, result)
		if n > 0:
			changed = True
			result = new_result

	return result, changed


def _apply_glossary_rules(text: str, glossary_rules):
	result = text
	changed = False
	for _, pattern, replacement in glossary_rules:
		new_result, n = pattern.subn(replacement, result)
		if n > 0:
			changed = True
			result = new_result
	return result, changed


def _translate_with_fallbacks(
	text: str,
	target_lang: str,
	trans_dict: dict[str, str],
	normalized_dict: dict[str, str],
	phrase_rules,
	glossary_rules,
	online_translator,
):
	processed_text = format_temperature(text)
	overrides = _get_hardcoded_overrides(target_lang)

	# 1) Exact match
	translated = trans_dict.get(processed_text)
	if translated is not None:
		return _finalize_translation(text, translated), "translated"

	# 2) Normalized exact match
	normalized_text = _normalize_text_for_lookup(processed_text)
	translated = normalized_dict.get(normalized_text)
	if translated is not None:
		return _finalize_translation(text, translated), "translated"

	override_translated = overrides.get(normalized_text.lower())
	if override_translated is not None:
		return _finalize_translation(text, override_translated), "translated"

	# 3) Line-by-line exact/normalized match (multiline cells)
	if "\n" in processed_text:
		lines = processed_text.splitlines()
		new_lines = []
		changed = False
		for line in lines:
			line_key = line.strip()
			line_translated = trans_dict.get(line_key)
			if line_translated is None:
				line_translated = normalized_dict.get(_normalize_text_for_lookup(line_key))
			if line_translated is not None:
				new_lines.append(line_translated)
				changed = True
			else:
				new_lines.append(line)
		if changed:
			return _finalize_translation(text, "\n".join(new_lines)), "translated"

	# 4) Conservative partial phrase replacement
	replaced_text = processed_text
	replaced = False
	for _, pattern, target in phrase_rules:
		if pattern.search(replaced_text):
			replaced_text = pattern.sub(target, replaced_text)
			replaced = True
	if replaced:
		return _finalize_translation(text, replaced_text), "translated"

	# 5) MSDS terminology fallback (when Master_DB has no similar phrase)
	templated_text, templated_changed = _translate_msds_templates(processed_text, target_lang)
	base_msds_text = templated_text if templated_changed else processed_text

	glossary_text, glossary_changed = _apply_glossary_rules(base_msds_text, glossary_rules)
	candidate_text = glossary_text if glossary_changed else base_msds_text
	candidate_changed = templated_changed or glossary_changed

	if candidate_changed and not _contains_english_text(candidate_text):
		return _finalize_translation(text, candidate_text), "translated"

	# 6) Precise online translation fallback (slow but high coverage)
	online_text, online_changed = _online_translate_precise(candidate_text, target_lang, online_translator)
	if online_changed:
		final_online = _finalize_translation(text, online_text)
		if not _contains_english_text(final_online):
			return final_online, "translated_online"
		residue_text, residue_changed = _translate_english_residue(
			final_online, target_lang, online_translator, glossary_rules
		)
		if residue_changed:
			return _finalize_translation(text, residue_text), "translated_online"
		return final_online, "translated_online"

	# Even when the online full-pass is not marked changed, try residue pass once.
	residue_text, residue_changed = _translate_english_residue(
		candidate_text, target_lang, online_translator, glossary_rules
	)
	if residue_changed:
		return _finalize_translation(text, residue_text), "translated_online"

	if _contains_english_text(candidate_text):
		progress_text = candidate_text
		for i in range(MAX_RESIDUE_TRANSLATION_PASSES):
			progress_text, changed = _translate_english_residue(
				progress_text,
				target_lang,
				online_translator,
				glossary_rules,
				delay_sec=1.8 + (0.4 * i),
			)
			if not changed or not _contains_english_text(progress_text):
				break
		if progress_text != candidate_text:
			return _finalize_translation(text, progress_text), "translated_online"

	if candidate_changed:
		return _finalize_translation(text, candidate_text), "translated"

	return _finalize_translation(text, processed_text), "unmatched"


# 4. 엑셀 파일 번역 처리 엔진 (서식 보존)
def process_excel(uploaded_file, target_lang, trans_dict):
	original_xlsx = uploaded_file.getvalue()
	wb = load_workbook(io.BytesIO(original_xlsx))
	normalized_dict = {_normalize_text_for_lookup(k): v for k, v in trans_dict.items()}
	phrase_rules = _build_phrase_rules(trans_dict)
	glossary_rules = _build_glossary_rules(target_lang)
	online_translator = _get_online_translator(target_lang)

	stats = {
		"translated": 0,
		"translated_online": 0,
		"kept_original": 0,
		"unmatched": 0,
	}

	for sheet_name in wb.sheetnames:
		ws = wb[sheet_name]
		for row in ws.iter_rows():
			for cell in row:
				if cell.value and isinstance(cell.value, str):
					original_text = cell.value.strip()
					if not original_text:
						continue

					processed_text = format_temperature(original_text)

					if should_force_keep_product_phrase(original_text):
						cell.value = _apply_mandatory_term_replacements(processed_text, target_lang)
						stats["kept_original"] += 1
					elif should_keep_original(processed_text):
						cell.value = _apply_mandatory_term_replacements(processed_text, target_lang)
						stats["kept_original"] += 1
					else:
						translated_text, status = _translate_with_fallbacks(
							original_text,
							target_lang,
							trans_dict,
							normalized_dict,
							phrase_rules,
							glossary_rules,
							online_translator,
						)
						if status == "translated":
							stats["translated"] += 1
						elif status == "translated_online":
							stats["translated_online"] += 1
						else:
							stats["unmatched"] += 1
						cell.value = _apply_mandatory_term_replacements(translated_text, target_lang)

	output = io.BytesIO()
	wb.save(output)
	translated_xlsx = output.getvalue()
	restored_xlsx = _restore_embedded_graphics(original_xlsx, translated_xlsx)
	final_output = io.BytesIO(restored_xlsx)
	final_output.seek(0)
	return final_output, stats


# 5. 앱 UI 구성 (메인 화면으로 번역란 이동)
st.markdown("### ⚙️ 1. 번역 언어 선택")
target_lang_display = st.radio(
	"어떤 언어로 번역하시겠습니까?",
	options=["베트남어 (Vietnamese)", "스페인어 (Spanish)", "중국어 (Chinese)"],
	horizontal=True,
)

lang_map = {
	"베트남어 (Vietnamese)": "Vietnamese",
	"스페인어 (Spanish)": "Spanish",
	"중국어 (Chinese)": "Chinese",
}
target_lang = lang_map[target_lang_display]
st.session_state["target_lang_runtime"] = target_lang

st.markdown("### 🔒 2. 비밀번호 인증")
entered_password = st.text_input("번역 실행 비밀번호를 입력하세요.", type="password")
is_authorized = entered_password == APP_PASSWORD

if not entered_password:
	st.info("비밀번호를 입력하면 업로드 및 번역 기능이 활성화됩니다.")
elif not is_authorized:
	st.error("비밀번호가 일치하지 않습니다.")

st.markdown("### 📊 3. 영문 MSDS 파일 업로드")
uploaded_file = st.file_uploader(
	"번역할 원본 엑셀 파일(_ENG.xlsx)을 올려주세요.",
	type=["xlsx"],
	disabled=not is_authorized,
)

if not is_authorized:
	st.warning("번역 기능은 비밀번호 인증 후 사용할 수 있습니다.")
	st.stop()

if uploaded_file is not None:
	original_filename = uploaded_file.name
	name_part, ext = os.path.splitext(original_filename)

	suffix_map = {
		"Vietnamese": "_VN",
		"Spanish": "_SPN",
		"Chinese": "_CHN",
	}
	target_suffix = suffix_map[target_lang]

	if "_ENG" in name_part:
		new_filename = name_part.replace("_ENG", target_suffix) + ext
	elif "_VN" in name_part or "_SPN" in name_part or "_CHN" in name_part:
		# 이미 다른 언어 꼬리표가 있으면 제거하고 새로 붙임
		new_filename = re.sub(r"_(VN|SPN|CHN)", target_suffix, name_part) + ext
	else:
		new_filename = name_part + target_suffix + ext

	st.info(f"선택 언어: **{target_lang_display}** | 예상 출력 파일명: **{new_filename}**")

	if st.button("🚀 번역 실행 및 엑셀 생성", use_container_width=True):
		with st.spinner(f"{target_lang_display}로 서식을 유지하며 번역 중입니다..."):
			trans_dict, source_logs = load_translation_dict(target_lang)

			if source_logs:
				st.caption(" / ".join(source_logs))

			if not trans_dict:
				st.warning("⚠️ 번역 사전이 비어있어 원문이 그대로 유지될 수 있습니다. DB 파일 형식을 확인해주세요.")

			processed_excel, stats = process_excel(uploaded_file, target_lang, trans_dict)

		st.success("✅ 번역 완료! 아래 버튼을 눌러 결과 파일을 다운로드하세요.")
		st.caption(
			f"DB/규칙 번역 {stats['translated']:,}개 / 정밀 온라인 번역 {stats['translated_online']:,}개 / 원문보호 {stats['kept_original']:,}개 / 미매칭 {stats['unmatched']:,}개"
		)

		st.download_button(
			label=f"📥 {new_filename} 다운로드",
			data=processed_excel,
			file_name=new_filename,
			mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
		)
